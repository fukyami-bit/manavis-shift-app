"""勤務希望Excel（文理ブロック分けのグリッド形式）を読み込むパーサー。

想定フォーマット:
- 1行目: タイトル
- 2行目: スタッフ列の上に「文」「理」ラベル
- 3行目: ヘッダー（不足 / 日付/曜日 / 平野 小西(職員) / 特記事項 / 開館時間 / スタッフ名... / 日付/曜日 ...）
- 4行目以降: 日付ごとのデータ行（開館時間・各スタッフの希望）

スタッフ名や人数は月ごとに変わりうるため、列名・文理ラベルから動的に列構成を検出する。
"""
from __future__ import annotations

import datetime
import re
import unicodedata
from dataclasses import dataclass

import openpyxl

from .models import DayInfo, RequestEntry, Staff

EXCLUDED_STAFF_NAMES = {"平野", "小西"}
DASH_CHARS = "ー‐－−~〜～"
UNAVAILABLE_TOKENS = {"×", "x", "X", "✕", "✖", "‪✕‬"}


def _normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKC", str(value))
    for ch in DASH_CHARS:
        text = text.replace(ch, "-")
    text = text.replace("　", " ").strip()
    return text


_RANGE_RE = re.compile(r"(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)")


def parse_cell(raw) -> RequestEntry | None:
    """1セルの値を解析する。呼び出し側で staff/date を付与する前提の中間結果を返す。"""
    if raw is None:
        return RequestEntry(staff="", date=None, raw="", type="none")

    if isinstance(raw, (datetime.datetime, datetime.date)):
        # Excelの自動日付変換の補正: 例えば "9-20" と入力すると 2026-09-20 に化ける。
        # 月=開始時刻, 日=終了時刻として復元する。
        start, end = float(raw.month), float(raw.day)
        return RequestEntry(
            staff="", date=None, raw=str(raw), type="range", start=start, end=end,
            note="Excelの日付自動変換を補正（元入力は恐らく数値の範囲）",
        )

    text = _normalize_text(raw)
    if text == "":
        return RequestEntry(staff="", date=None, raw=text, type="none")

    if text in UNAVAILABLE_TOKENS:
        return RequestEntry(staff="", date=None, raw=text, type="unavailable")

    if text == "終日":
        return RequestEntry(staff="", date=None, raw=text, type="full_day")

    tentative = False
    body = text
    if body.startswith("△"):
        tentative = True
        body = body[1:].strip()
        body = body.strip("()（）")

    if body == "" or body == "終日":
        t = "full_day" if tentative else "none"
        return RequestEntry(staff="", date=None, raw=text, type=t if body else "full_day", tentative=tentative)

    m = _RANGE_RE.search(body)
    if m:
        start, end = float(m.group(1)), float(m.group(2))
        note = None
        if body[m.start():m.end()] != body:
            note = f"元の記載: {text}"
        return RequestEntry(
            staff="", date=None, raw=text, type="range", start=start, end=end,
            tentative=tentative, note=note,
        )

    return RequestEntry(staff="", date=None, raw=text, type="unknown", tentative=tentative, note="形式を認識できませんでした。手動で確認してください")


@dataclass
class SheetLayout:
    """出力時に元ファイルへ書き戻すためのセル座標情報"""
    sheet_name: str
    header_row: int
    category_row: int
    shortage_col: int
    primary_date_col: int
    weekday_col: int
    note_col: int
    open_hours_col: int
    staff_columns: list  # list[(col_idx, name, category, wage)]
    data_start_row: int
    data_end_row: int


@dataclass
class ParsedSheet:
    staff: list  # list[Staff]
    days: list  # list[DayInfo]
    requests: list  # list[RequestEntry]
    warnings: list  # list[str]
    layout: "SheetLayout" = None


def parse_workbook(file_obj, hourly_wage_overrides: dict | None = None, default_wage: int = 1200) -> ParsedSheet:
    """アップロードされたExcelファイル（file-likeまたはパス）を解析する。"""
    hourly_wage_overrides = hourly_wage_overrides or {}
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row_idx = None
    for r in range(1, 8):
        val = ws.cell(row=r, column=1).value
        if val is not None and _normalize_text(val) == "不足":
            header_row_idx = r
            break
    if header_row_idx is None:
        raise ValueError("ヘッダー行（1列目に「不足」）が見つかりませんでした。フォーマットを確認してください。")

    category_row_idx = header_row_idx - 1
    max_col = ws.max_column

    staff_columns = []  # list[(col_idx, name, category)]
    open_hours_col = None
    note_col = None
    date_columns = []  # list of column idx where header == 日付/曜日

    for c in range(1, max_col + 1):
        header_val = ws.cell(row=header_row_idx, column=c).value
        if header_val is None:
            continue
        header_text = _normalize_text(header_val)
        if header_text == "日付/曜日":
            date_columns.append(c)
        elif header_text == "開館時間":
            open_hours_col = c
        elif header_text == "特記事項":
            note_col = c
            continue
        elif header_text == "不足":
            continue
        else:
            name = header_text
            if name in EXCLUDED_STAFF_NAMES:
                continue
            category_val = ws.cell(row=category_row_idx, column=c).value
            category = _normalize_text(category_val) if category_val else None
            if category in ("文", "理"):
                wage = hourly_wage_overrides.get(name, default_wage)
                staff_columns.append((c, name, category, wage))

    if not date_columns:
        raise ValueError("日付/曜日列が見つかりませんでした。")
    if open_hours_col is None:
        raise ValueError("開館時間列が見つかりませんでした。")

    primary_date_col = date_columns[0]
    weekday_col = primary_date_col + 1

    staff_list = [Staff(name=n, category=cat, hourly_wage=w) for _, n, cat, w in staff_columns]

    days: list[DayInfo] = []
    requests: list[RequestEntry] = []
    warnings: list[str] = []
    first_data_row = None

    r = header_row_idx + 1
    while r <= ws.max_row:
        date_val = ws.cell(row=r, column=primary_date_col).value
        if not isinstance(date_val, (datetime.datetime, datetime.date)):
            if days:
                break
            r += 1
            if r - header_row_idx > 15:
                break
            continue

        if first_data_row is None:
            first_data_row = r

        d = date_val.date() if isinstance(date_val, datetime.datetime) else date_val
        weekday_val = ws.cell(row=r, column=weekday_col).value
        weekday = _normalize_text(weekday_val) if weekday_val else ""

        open_raw = ws.cell(row=r, column=open_hours_col).value
        open_entry = parse_cell(open_raw)
        open_start, open_end = None, None
        if open_entry.type == "range":
            open_start, open_end = open_entry.start, open_entry.end
        elif open_raw is not None:
            warnings.append(f"{d} の開館時間「{open_raw}」を解析できませんでした")

        special_note = None
        note_col_effective = note_col or (primary_date_col + 3)
        note_val = ws.cell(row=r, column=note_col_effective).value
        if note_val:
            special_note = _normalize_text(note_val)

        days.append(DayInfo(date=d, weekday=weekday, open_start=open_start, open_end=open_end, special_note=special_note))

        for col_idx, name, _cat, _wage in staff_columns:
            raw_val = ws.cell(row=r, column=col_idx).value
            entry = parse_cell(raw_val)
            entry.staff = name
            entry.date = d
            if entry.type == "unknown":
                warnings.append(f"{d} {name}: 「{entry.raw}」の形式を認識できませんでした（手動確認推奨）")
            if entry.type != "none":
                requests.append(entry)

        r += 1

    layout = SheetLayout(
        sheet_name=ws.title,
        header_row=header_row_idx,
        category_row=category_row_idx,
        shortage_col=1,
        primary_date_col=primary_date_col,
        weekday_col=weekday_col,
        note_col=note_col or (primary_date_col + 3),
        open_hours_col=open_hours_col,
        staff_columns=staff_columns,
        data_start_row=first_data_row if first_data_row is not None else header_row_idx + 1,
        data_end_row=r - 1,
    )

    return ParsedSheet(staff=staff_list, days=days, requests=requests, warnings=warnings, layout=layout)
