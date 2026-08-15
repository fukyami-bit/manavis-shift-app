"""確定シフトをExcelに書き出す。

2シート構成:
- 「出勤体制」: 元の勤務希望シートと同じ雰囲気のグリッド形式（文理ブロック分け・開館時間・不足列）
- 「時間帯一覧」: 日付ごとに「誰が何時から何時まで」を時系列で並べた読みやすいリスト
"""
from __future__ import annotations

from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Assignment, DayInfo, ScheduleResult, Staff

HEADER_FILL = PatternFill("solid", fgColor="F1EFE8")
SHORTAGE_FILL = PatternFill("solid", fgColor="FAEEDA")
BOLD = Font(bold=True)


def format_time(hours: float) -> str:
    total_minutes = round(hours * 60)
    hh, mm = divmod(total_minutes, 60)
    return f"{hh}:{mm:02d}"


def build_workbook(
    staff_list: list[Staff],
    days: list[DayInfo],
    result: ScheduleResult,
    month_label: str = "",
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    _build_grid_sheet(wb.active, staff_list, days, result, month_label)
    ws2 = wb.create_sheet("時間帯一覧")
    _build_timeline_sheet(ws2, days, result)
    return wb


def _build_grid_sheet(ws, staff_list, days, result: ScheduleResult, month_label: str):
    ws.title = "出勤体制"
    bunkei = [s for s in staff_list if s.category == "文"]
    rikei = [s for s in staff_list if s.category == "理"]
    ordered_staff = bunkei + rikei

    by_day = defaultdict(list)
    for a in result.assignments:
        by_day[a.date].append(a)

    shortage_by_day = defaultdict(list)
    for sh in result.shortages:
        shortage_by_day[sh["date"]].append(sh)

    ws.cell(row=1, column=1, value=f"{month_label} 出勤体制（確定）").font = Font(bold=True, size=13)

    col_names = ["不足", "日付", "曜日", "特記事項", "開館時間"] + [s.name for s in ordered_staff]
    header_row = 3
    for c, name in enumerate(col_names, start=1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.font = BOLD
        cell.fill = HEADER_FILL

    cat_row = header_row - 1
    offset = 6  # スタッフ列の開始列
    for i, s in enumerate(ordered_staff):
        ws.cell(row=cat_row, column=offset + i, value=s.category)

    r = header_row + 1
    for day in days:
        assigns = sorted(by_day.get(day.date, []), key=lambda a: a.start)
        shortages_today = shortage_by_day.get(day.date, [])

        ws.cell(row=r, column=1, value="; ".join(f"{s['band']} {s['available']}/{s['required']}" for s in shortages_today) or None)
        ws.cell(row=r, column=2, value=day.date.strftime("%m/%d"))
        ws.cell(row=r, column=3, value=day.weekday)
        ws.cell(row=r, column=4, value=day.special_note)
        if day.open_start is not None:
            ws.cell(row=r, column=5, value=f"{day.open_start:g}-{day.open_end:g}")

        assigns_by_staff = {a.staff: a for a in assigns}
        for i, s in enumerate(ordered_staff):
            a = assigns_by_staff.get(s.name)
            if a:
                text = f"{format_time(a.start)}-{format_time(a.end)}（{a.hours:g}h）"
                if a.tentative:
                    text = "△" + text
                ws.cell(row=r, column=offset + i, value=text)

        if shortages_today:
            for c in range(1, offset + len(ordered_staff)):
                ws.cell(row=r, column=c).fill = SHORTAGE_FILL
        r += 1

    # 合計行
    total_row = r + 1
    ws.cell(row=total_row, column=3, value="合計時間 / 人件費").font = BOLD
    for i, s in enumerate(ordered_staff):
        stat = result.staff_stats.get(s.name, {})
        total_hours = sum(a.hours for a in result.assignments if a.staff == s.name)
        cost = total_hours * s.hourly_wage
        ratio = stat.get("ratio")
        ratio_text = f"{ratio*100:.0f}%" if ratio is not None else "-"
        ws.cell(row=total_row, column=offset + i, value=f"{total_hours:g}h / {int(cost):,}円 / 希望充足{ratio_text}")

    ws.cell(row=total_row + 2, column=1, value=f"合計人件費: {int(result.total_cost):,}円 / 予算: {result.budget:,}円").font = BOLD

    for c in range(1, offset + len(ordered_staff)):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.column_dimensions[get_column_letter(4)].width = 20


def build_template_workbook(
    original_file_obj,
    layout,
    staff_list: list[Staff],
    days: list[DayInfo],
    result: ScheduleResult,
) -> openpyxl.Workbook:
    """アップロードされた元ファイルそのものをテンプレートとして使い、
    希望が入っていたセルを確定シフトで上書きする。フォント・罫線・列幅・
    色などの元の書式は変更しない。"""
    wb = openpyxl.load_workbook(original_file_obj)
    ws = wb[layout.sheet_name]

    by_day_staff = {}
    for a in result.assignments:
        by_day_staff[(a.date, a.staff)] = a

    shortage_by_day = defaultdict(list)
    for sh in result.shortages:
        shortage_by_day[sh["date"]].append(sh)

    for row, day in zip(range(layout.data_start_row, layout.data_end_row + 1), days):
        shortages_today = shortage_by_day.get(day.date, [])
        shortage_cell = ws.cell(row=row, column=layout.shortage_col)
        shortage_cell.value = (
            "; ".join(f"{s['band']} {s['available']}/{s['required']}" for s in shortages_today) or None
        )
        if shortages_today:
            shortage_cell.fill = SHORTAGE_FILL

        for col_idx, name, _cat, _wage in layout.staff_columns:
            cell = ws.cell(row=row, column=col_idx)
            a = by_day_staff.get((day.date, name))
            if a is None:
                cell.value = None
                continue
            text = f"{a.start:g}-{a.end:g}"
            if a.tentative:
                text = "△" + text
            cell.value = text

    ws2 = wb.create_sheet("時間帯一覧")
    _build_timeline_sheet(ws2, days, result)
    return wb


def _build_timeline_sheet(ws, days, result: ScheduleResult):
    headers = ["日付", "曜日", "時間帯", "担当"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = BOLD
        cell.fill = HEADER_FILL

    by_day = defaultdict(list)
    for a in result.assignments:
        by_day[a.date].append(a)

    r = 2
    for day in days:
        assigns = sorted(by_day.get(day.date, []), key=lambda a: a.start)
        if not assigns:
            continue
        for a in assigns:
            ws.cell(row=r, column=1, value=day.date.strftime("%m/%d"))
            ws.cell(row=r, column=2, value=day.weekday)
            time_text = f"{format_time(a.start)} - {format_time(a.end)}"
            if a.tentative:
                time_text += "（仮）"
            ws.cell(row=r, column=3, value=time_text)
            ws.cell(row=r, column=4, value=a.staff)
            r += 1

    for c, w in zip(range(1, 5), (10, 8, 20, 14)):
        ws.column_dimensions[get_column_letter(c)].width = w
